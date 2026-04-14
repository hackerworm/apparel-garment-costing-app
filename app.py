from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os, re, fitz, pdfplumber, pandas as pd
import base64, tempfile, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

# ── PDF extraction ─────────────────────────────────────────────────────────────
FRONT_RECT = fitz.Rect(62.846, 143.286, 210.034, 343.286)
ZOOM       = 2

# ── Brand palette (ARGB, no #) ────────────────────────────────────────────────
COGNAC      = "6B3A2A"
COGNAC_DARK = "4E2A1E"
CREAM       = "F5EFE6"
CREAM_DARK  = "EDE3D6"
WHITE       = "FFFFFF"
MUTED       = "7A6A62"
LIGHT_ROW   = "FDFAF7"

def sd(color="D9C9C0", style="thin"):
    return Side(style=style, color=color)

def bdr_bottom(color="D9C9C0", style="thin"):
    return Border(bottom=sd(color, style))

def bdr_full(color=COGNAC, style="thin"):
    s = sd(color, style)
    return Border(top=s, bottom=s, left=s, right=s)

def extract_image_b64(page, rect):
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(ZOOM, ZOOM), clip=rect)
        return base64.b64encode(pix.tobytes("png")).decode()
    except:
        return None

def extract_style_number(filename):
    m = re.search(r"(OB\d{7})", filename, re.IGNORECASE)
    return m.group(1) if m else os.path.splitext(filename)[0]

def flt(v):
    try: return float(v or 0)
    except: return 0.0

# ── Excel generation ───────────────────────────────────────────────────────────
def generate_excel(data):
    hdr      = data.get("header", {})
    secs     = data.get("sections", {})
    enabled  = data.get("enabled", {})
    totals   = data.get("totals", {})
    ovhd_pct = flt(data.get("overheadPct", 0))
    mrgn_pct = flt(data.get("marginPct",   0))
    ex_rate  = flt(hdr.get("exRate", 0))
    img_b64  = data.get("frontImageB64")

    keys     = ["fabricShell","lining","trims","embellishment","cmt",
                "cutRecut","washing","labels","packing","courierTest"]
    active   = lambda k: flt(totals.get(k, 0)) if enabled.get(k) else 0.0
    subtotal = sum(active(k) for k in keys)
    overhead = subtotal * ovhd_pct / 100
    mrg_base = subtotal + overhead
    margin   = mrg_base * mrgn_pct / 100
    total    = mrg_base + margin
    total_usd = total / ex_rate if ex_rate > 0 else None

    INR = '[$₹-hi-IN]#,##0.00'
    USD = '"$"#,##0.00'

    wb  = Workbook()

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — DETAILED COSTING
    # Columns: A(desc,28) B(val1,14) C(val2,14) D(val3,14) E(amt,14) | F(img,16)
    # Total width ~84 chars — fits comfortably on A4 landscape
    # ══════════════════════════════════════════════════════════════
    ws  = wb.active
    ws.title = "Detailed Costing"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation   = "landscape"
    ws.page_setup.fitToPage     = True
    ws.page_setup.fitToWidth    = 1

    # Column widths — tighter than before
    COL_W = {"A": 26, "B": 13, "C": 13, "D": 13, "E": 14, "F": 16}
    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    row = 1

    # ── Cell writer helper ───────────────────────────────────────
    def wc(ref, val=None, bold=False, sz=9, fg=WHITE, bg=None,
           ha="left", va="center", wrap=False, bdr=None,
           italic=False, fmt=None, indent=0):
        c = ws[ref]
        if val is not None:
            c.value = val
        c.font      = Font(name="Calibri", bold=bold, size=sz, color=fg, italic=italic)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap, indent=indent)
        if bg:  c.fill   = PatternFill("solid", fgColor=bg)
        if bdr: c.border = bdr
        if fmt: c.number_format = fmt
        return c

    def spacer(h=5):
        nonlocal row
        ws.row_dimensions[row].height = h
        row += 1

    def merge(start_col, end_col, r=None):
        r = r or row
        ws.merge_cells(f"{start_col}{r}:{end_col}{r}")

    # ── BANNER ───────────────────────────────────────────────────
    ws.row_dimensions[row].height = 40
    merge("A","E")
    wc(f"A{row}", "CONTEMPORARY CLASSIQUE LLP",
       bold=True, sz=16, fg=WHITE, bg=COGNAC_DARK, indent=1)
    # Image column background
    ws[f"F{row}"].fill = PatternFill("solid", fgColor=COGNAC_DARK)
    row += 1

    ws.row_dimensions[row].height = 16
    merge("A","E")
    wc(f"A{row}", "PROTO SAMPLE COSTING SHEET",
       sz=8, fg="C4A090", bg=COGNAC_DARK, italic=True, indent=1)
    ws[f"F{row}"].fill = PatternFill("solid", fgColor=COGNAC_DARK)
    row += 1

    spacer(3)

    # ── STYLE INFO BLOCK + IMAGE ─────────────────────────────────
    # Layout: A=label(narrow) | B-C=value(merged) | D-E=label+value | F=image
    info_start = row
    fields_left = [
        ("STYLE NO.",  hdr.get("styleNo","")),
        ("STYLE NAME", hdr.get("styleName","")),
        ("BUYER",      hdr.get("buyer","")),
        ("SEASON",     hdr.get("season","")),
    ]
    fields_right = [
        ("CATEGORY",      hdr.get("garmentCategory","")),
        ("DATE",          hdr.get("date","")),
        ("TOTAL PIECES",  str(hdr.get("totalPieces","")) or "—"),
        ("",              ""),   # padding row to match left
    ]

    for (ll, lv), (rl, rv) in zip(fields_left, fields_right):
        ws.row_dimensions[row].height = 18
        b = bdr_bottom()

        # Left pair: A=label, B-C=value
        wc(f"A{row}", ll, bold=True, sz=8, fg=MUTED, bg=CREAM, bdr=b)
        merge("B","C")
        wc(f"B{row}", lv, sz=9, fg="1C1C1C", bg=WHITE, bdr=b)
        ws[f"C{row}"].border = b

        # Right pair: D=label, E=value
        wc(f"D{row}", rl, bold=True, sz=8, fg=MUTED, bg=CREAM, bdr=b)
        wc(f"E{row}", rv, sz=9, fg="1C1C1C", bg=WHITE, bdr=b)

        # Image column
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    info_end = row - 1

    # Embed front image in column F
    if img_b64:
        try:
            pil   = PILImage.open(io.BytesIO(base64.b64decode(img_b64)))
            n_rows = info_end - info_start + 1
            max_h  = int(n_rows * 18 * 1.33)   # approx pixel height of rows
            ratio  = max_h / pil.height
            new_w  = int(pil.width * ratio)
            pil    = pil.resize((new_w, max_h), PILImage.LANCZOS)
            buf2   = io.BytesIO()
            pil.save(buf2, "PNG")
            buf2.seek(0)
            xl_img        = XLImage(buf2)
            xl_img.width  = new_w
            xl_img.height = max_h
            ws.column_dimensions["F"].width = max(16, int(new_w / 7.5) + 1)
            ws.add_image(xl_img, f"F{info_start}")
        except Exception as e:
            print(f"Image embed error: {e}")

    spacer(6)

    # ── SECTION HELPERS ──────────────────────────────────────────
    def sec_header(title):
        nonlocal row
        ws.row_dimensions[row].height = 22
        merge("A","E")
        wc(f"A{row}", f"  {title}", bold=True, sz=9, fg=WHITE, bg=COGNAC)
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    def col_hdr(*labels):
        nonlocal row
        ws.row_dimensions[row].height = 15
        cols = "ABCDE"
        for i, lbl in enumerate(labels):
            if i >= len(cols): break
            c = ws[f"{cols[i]}{row}"]
            c.value     = lbl
            c.font      = Font(name="Calibri", bold=True, size=8, color=COGNAC)
            c.fill      = PatternFill("solid", fgColor=CREAM_DARK)
            c.alignment = Alignment(horizontal="right" if i > 0 else "left", vertical="center")
            c.border    = Border(bottom=sd(COGNAC, "medium"))
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    def data_r(*vals, alt=False):
        nonlocal row
        ws.row_dimensions[row].height = 15
        bg   = LIGHT_ROW if alt else WHITE
        cols = "ABCDE"
        for i, v in enumerate(vals):
            if i >= len(cols): break
            c      = ws[f"{cols[i]}{row}"]
            c.value = v
            is_n   = isinstance(v, float)
            c.font  = Font(name="Calibri", size=9, color="1C1C1C")
            c.fill  = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="right" if is_n else "left", vertical="center")
            c.border = bdr_bottom()
            if is_n: c.number_format = INR
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    def inline_total(label, value):
        """Single line: label on left, value on right — no repeated heading."""
        nonlocal row
        ws.row_dimensions[row].height = 16
        merge("A","D")
        wc(f"A{row}", f"  {label}", bold=True, sz=9,
           fg=COGNAC, bg=CREAM, italic=True,
           bdr=Border(top=sd(COGNAC), bottom=sd(COGNAC)))
        merge("E","E")
        c = wc(f"E{row}", value, bold=True, sz=9, fg=COGNAC, bg=CREAM, ha="right",
               bdr=Border(top=sd(COGNAC), bottom=sd(COGNAC)))
        c.number_format = INR
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    def single_line_section(key, label, rate_val):
        """
        For simple single-rate sections: one cognac bar with label on left,
        rate in middle, total on right. No repeated sub-heading.
        """
        if not enabled.get(key): return
        nonlocal row
        ws.row_dimensions[row].height = 18

        # A-B: section label
        merge("A","B")
        wc(f"A{row}", f"  {label}", bold=True, sz=9, fg=WHITE, bg=COGNAC)

        # C: sub-label "Rate / pc"
        wc(f"C{row}", "Rate / pc", bold=True, sz=8, fg="C4A090", bg=COGNAC, ha="right")

        # D: rate value
        c = wc(f"D{row}", flt(rate_val), bold=True, sz=9, fg=WHITE, bg=COGNAC, ha="right")
        c.number_format = INR

        # E: total (same as rate for single-rate sections)
        c2 = wc(f"E{row}", active(key), bold=True, sz=9, fg=WHITE, bg=COGNAC, ha="right")
        c2.number_format = INR

        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1
        spacer(3)

    def simple_section(key, title, rows_data):
        """Multi-row simple section (description + rate)."""
        if not enabled.get(key): return
        sec_header(title)
        col_hdr("Description", "Rate / pc (₹)")
        for i, r in enumerate(rows_data or []):
            data_r(r.get("description",""), flt(r.get("rate",0)), alt=i%2==1)
        inline_total(f"{title} Total", active(key))
        spacer(3)

    # ── FABRIC SHELL ─────────────────────────────────────────────
    if enabled.get("fabricShell"):
        sec_header("FABRIC — SHELL")
        col_hdr("Fabric Name", "Width", "Consumption / pc", "Rate / Mtr (₹)", "Cost (₹)")
        for i, r in enumerate(secs.get("fabricShell",[])):
            data_r(r.get("name",""), r.get("width",""),
                   r.get("consumption",""),
                   flt(r.get("ratePerMeter",0)),
                   flt(r.get("singlePcCost",0)), alt=i%2==1)
        inline_total("Fabric Shell Total", active("fabricShell"))
        spacer(3)

    # ── LINING ───────────────────────────────────────────────────
    if enabled.get("lining"):
        sec_header("FABRIC — LINING")
        col_hdr("Fabric Name", "Width", "Consumption / pc", "Rate / Mtr (₹)", "Cost (₹)")
        for i, r in enumerate(secs.get("lining",[])):
            data_r(r.get("name",""), r.get("width",""),
                   r.get("consumption",""),
                   flt(r.get("ratePerMeter",0)),
                   flt(r.get("singlePcCost",0)), alt=i%2==1)
        inline_total("Lining Total", active("lining"))
        spacer(3)

    # ── TRIMS & ACCESSORIES ──────────────────────────────────────
    # Columns: Description | Qty | Unit | Rate/Unit | Cost
    if enabled.get("trims"):
        sec_header("TRIMS & ACCESSORIES")
        col_hdr("Description", "Qty / Avg", "Unit", "Rate / Unit (₹)", "Cost (₹)")
        for i, r in enumerate(secs.get("trims",[])):
            data_r(r.get("description",""),
                   r.get("qty",""),
                   r.get("unit",""),          # unit: pcs / mtr / yd / set etc.
                   flt(r.get("rate",0)),
                   flt(r.get("cost",0)), alt=i%2==1)
        inline_total("Trims Total", active("trims"))
        spacer(3)

    # ── EMBELLISHMENT ────────────────────────────────────────────
    if enabled.get("embellishment"):
        simple_section("embellishment", "EMBELLISHMENT / EMBROIDERY",
                       secs.get("embellishment",[]))

    # ── CMT ─────────────────────────────────────────────────────
    single_line_section("cmt", "CMT  (Cutting, Making & Trimming)",
                        secs.get("cmt", 0))

    # ── CUT & RECUT ──────────────────────────────────────────────
    single_line_section("cutRecut", "CUT & RECUT",
                        secs.get("cutRecut", 0))

    # ── WASHING / FINISHING ──────────────────────────────────────
    # Single-line if only one entry with same name as section, else table
    washing_rows = secs.get("washing", [])
    if enabled.get("washing"):
        if len(washing_rows) == 1:
            single_line_section("washing",
                                f"WASHING / FINISHING — {washing_rows[0].get('description','')}",
                                washing_rows[0].get("rate", 0))
        else:
            simple_section("washing", "WASHING / FINISHING", washing_rows)

    # ── LABELS & TAGS ─────────────────────────────────────────────
    labels_rows = secs.get("labels", [])
    if enabled.get("labels"):
        if len(labels_rows) == 1:
            single_line_section("labels",
                                f"LABELS & TAGS — {labels_rows[0].get('description','')}",
                                labels_rows[0].get("rate", 0))
        else:
            simple_section("labels", "LABELS & TAGS", labels_rows)

    # ── PACKING & MATERIAL ────────────────────────────────────────
    packing_rows = secs.get("packing", [])
    if enabled.get("packing"):
        if len(packing_rows) == 1:
            single_line_section("packing",
                                f"PACKING & MATERIAL — {packing_rows[0].get('description','')}",
                                packing_rows[0].get("rate", 0))
        else:
            simple_section("packing", "PACKING & MATERIAL", packing_rows)

    # ── COURIER & TESTING ─────────────────────────────────────────
    single_line_section("courierTest", "COURIER & TESTING",
                        secs.get("courierTest", 0))

    spacer(4)

    # ── GRAND TOTAL BLOCK ─────────────────────────────────────────
    def total_row(label, value, highlight=False, usd=False):
        nonlocal row
        ws.row_dimensions[row].height = 20 if highlight else 16
        bg   = COGNAC if highlight else CREAM
        fg_c = WHITE  if highlight else COGNAC
        sz   = 11 if highlight else 9
        bdr  = bdr_full(COGNAC, "medium") if highlight else bdr_bottom(COGNAC)

        merge("A","D")
        wc(f"A{row}", f"  {label}", bold=True, sz=sz,
           fg=fg_c, bg=bg, bdr=bdr, indent=1)
        merge("E","E")
        c = wc(f"E{row}", value, bold=True, sz=sz,
               fg=fg_c, bg=bg, ha="right", bdr=bdr)
        c.number_format = USD if usd else INR
        ws[f"F{row}"].fill = PatternFill("solid", fgColor=CREAM)
        row += 1

    total_row("Subtotal",                          subtotal)
    total_row(f"Overhead  ({ovhd_pct:.1f}%)",      overhead)
    total_row(f"Profit Margin  ({mrgn_pct:.1f}%)", margin)
    spacer(3)
    total_row("TOTAL COST  (INR)",                 total, highlight=True)

    if total_usd is not None:
        spacer(3)
        total_row(f"TOTAL COST  (USD @ ₹{ex_rate:.2f} = 1 USD)",
                  total_usd, highlight=True, usd=True)

    spacer(10)

    # Footer
    ws.row_dimensions[row].height = 13
    merge("A","E")
    wc(f"A{row}",
       "Contemporary Classique LLP  |  Proto Sample Costing  |  Confidential",
       sz=7, fg=MUTED, bg=CREAM, italic=True, ha="center")

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — SUMMARY
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 18

    r = 1

    def w2(ref, val=None, bold=False, sz=9, fg=WHITE, bg=None,
           ha="left", bdr=None, italic=False, fmt=None):
        c = ws2[ref]
        if val is not None: c.value = val
        c.font      = Font(name="Calibri", bold=bold, size=sz, color=fg, italic=italic)
        c.alignment = Alignment(horizontal=ha, vertical="center")
        if bg:  c.fill   = PatternFill("solid", fgColor=bg)
        if bdr: c.border = bdr
        if fmt: c.number_format = fmt
        return c

    def s2sp(h=6):
        nonlocal r
        ws2.row_dimensions[r].height = h
        r += 1

    def s2merge(r_=None):
        nonlocal r
        ws2.merge_cells(f"A{r_ or r}:B{r_ or r}")

    # Banner
    ws2.row_dimensions[r].height = 38
    s2merge(r)
    w2(f"A{r}", "CONTEMPORARY CLASSIQUE LLP",
       bold=True, sz=15, fg=WHITE, bg=COGNAC_DARK)
    r += 1
    ws2.row_dimensions[r].height = 15
    s2merge(r)
    w2(f"A{r}", "PROTO SAMPLE — COST SUMMARY",
       sz=8, fg="C4A090", bg=COGNAC_DARK, italic=True)
    r += 1
    s2sp()

    # Info
    for label, value in [
        ("STYLE NO.",    hdr.get("styleNo","")),
        ("STYLE NAME",   hdr.get("styleName","")),
        ("BUYER",        hdr.get("buyer","")),
        ("SEASON",       hdr.get("season","")),
        ("DATE",         hdr.get("date","")),
        ("TOTAL PIECES", str(hdr.get("totalPieces","")) or "—"),
    ]:
        ws2.row_dimensions[r].height = 17
        b = bdr_bottom()
        w2(f"A{r}", label, bold=True, sz=8, fg=MUTED, bg=CREAM, bdr=b)
        w2(f"B{r}", value, sz=9,  fg="1C1C1C", bdr=b)
        r += 1

    s2sp()

    # Breakdown header
    ws2.row_dimensions[r].height = 20
    w2(f"A{r}", "  COST BREAKDOWN", bold=True, sz=9, fg=WHITE, bg=COGNAC)
    w2(f"B{r}", "AMOUNT (INR)",     bold=True, sz=9, fg=WHITE, bg=COGNAC, ha="right")
    r += 1

    sec_map = [
        ("Fabric (Shell)",      "fabricShell"),
        ("Fabric (Lining)",     "lining"),
        ("Trims & Accessories", "trims"),
        ("Embellishment",       "embellishment"),
        ("CMT",                 "cmt"),
        ("Cut & Recut",         "cutRecut"),
        ("Washing / Finishing", "washing"),
        ("Labels & Tags",       "labels"),
        ("Packing & Material",  "packing"),
        ("Courier & Testing",   "courierTest"),
    ]
    active_secs = [(l, k) for l, k in sec_map if enabled.get(k)]
    for i, (label, key) in enumerate(active_secs):
        ws2.row_dimensions[r].height = 17
        bg  = LIGHT_ROW if i%2==1 else WHITE
        b   = bdr_bottom()
        w2(f"A{r}", label,       sz=9, fg="1C1C1C", bg=bg, bdr=b)
        c = w2(f"B{r}", active(key), sz=9, fg="1C1C1C", bg=bg, ha="right", bdr=b, fmt=INR)
        r += 1

    s2sp(3)
    for label, value in [
        (f"Overhead ({ovhd_pct:.1f}%)",    overhead),
        (f"Profit Margin ({mrgn_pct:.1f}%)", margin),
    ]:
        ws2.row_dimensions[r].height = 16
        b = bdr_bottom()
        w2(f"A{r}", label, sz=9, fg=MUTED, bg=WHITE, bdr=b)
        w2(f"B{r}", value, sz=9, fg=MUTED, bg=WHITE, ha="right", bdr=b, fmt=INR)
        r += 1

    s2sp(5)

    def s2_grand(label, value, usd=False):
        nonlocal r
        ws2.row_dimensions[r].height = 26
        b = bdr_full(COGNAC, "medium")
        w2(f"A{r}", f"  {label}", bold=True, sz=12, fg=WHITE, bg=COGNAC, bdr=b)
        w2(f"B{r}", value, bold=True, sz=12, fg=WHITE, bg=COGNAC,
           ha="right", bdr=b, fmt=USD if usd else INR)
        r += 1

    s2_grand("TOTAL COST  (INR)", total)
    if total_usd is not None:
        s2sp(3)
        s2_grand(f"TOTAL COST  (USD @ ₹{ex_rate:.2f})", total_usd, usd=True)

    s2sp(12)
    ws2.row_dimensions[r].height = 13
    s2merge(r)
    w2(f"A{r}", "Contemporary Classique LLP  |  Confidential",
       sz=7, fg=MUTED, bg=CREAM, italic=True, ha="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Frontend static file serving ───────────────────────────────────────────────
# In production (packaged EXE) Flask serves the React build from a 'frontend'
# folder sitting next to the executable.  In dev, Vite serves separately.
import sys

def get_frontend_folder():
    if getattr(sys, 'frozen', False):
        # PyInstaller sets sys._MEIPASS when running as bundled exe
        base = sys._MEIPASS
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, 'frontend_build')

FRONTEND = get_frontend_folder()

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_frontend(path):
    if not os.path.isdir(FRONTEND):
        return jsonify({"error": "Frontend build not found"}), 404
    # Serve actual file if it exists, else fallback to index.html (SPA routing)
    target = os.path.join(FRONTEND, path)
    if path and os.path.isfile(target):
        return send_file(target)
    return send_file(os.path.join(FRONTEND, 'index.html'))

# ── API Routes ─────────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})

@app.route("/extract", methods=["POST"])
def extract():
    """Extract style number + front image only. No BOM import."""
    if "pdf" not in request.files:
        return jsonify({"error": "No PDF provided"}), 400
    file = request.files["pdf"]
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name
    try:
        style_number = extract_style_number(file.filename or "techpack.pdf")
        doc   = fitz.open(tmp_path)
        front = extract_image_b64(doc[0], FRONT_RECT)
        doc.close()
        # Return style number + image only — no BOM
        return jsonify({
            "style_number": style_number,
            "front_image":  front,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        os.unlink(tmp_path)


@app.route("/export-excel", methods=["POST"])
def export_excel():
    try:
        data  = request.get_json(force=True)
        buf   = generate_excel(data)
        style = data.get("header", {}).get("styleNo", "proto")
        fname = f"{style}_costing_CC.xlsx"
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    is_dev = not getattr(sys, 'frozen', False)
    print("Contemporary Classique — Proto Costing Backend")
    print("Running on http://localhost:5000")
    app.run(debug=is_dev, port=5000, use_reloader=is_dev)

# Override __main__ block for production
